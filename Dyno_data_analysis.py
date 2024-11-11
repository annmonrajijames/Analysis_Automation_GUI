import pandas as pd
import openpyxl
import os

# Load the CSV file
input_file_path = r"C:\Users\kamalesh.kb\NOV_7_IDC_DATA\Cleaned_data\Modified_FL-07_11_2024-NDURO_190KG_ORD_TRIAL-19_MORNING_TRIAL_DATA.csv"
input_df = pd.read_csv(input_file_path)

# Filter data based on SOC conditions
df = input_df[(input_df['SOC'] < 19) & (input_df['SOC'] > 2)]

# Calculate the mean, max, and min of 'PackCurr' column
pack_curr_mean = abs(df['PackCurr'].mean())
pack_curr_max = df['PackCurr'].max()
pack_curr_min = df['PackCurr'].min()

# Calculate the mean, max, and min of 'ForceAtWheel' column
force_mean = df['ForceAtWheel'].mean()
force_max = df['ForceAtWheel'].max()
force_min = df['ForceAtWheel'].min()

# Parse the 'Date Time' column into datetime format with millisecond precision
df['Date Time'] = pd.to_datetime(df['Date Time'], format='%d:%m:%Y %H:%M:%S:%f')

# Calculate time difference in seconds
df['localtime_Diff'] = df['Date Time'].diff().dt.total_seconds().fillna(0)

# Calculate the instantaneous power and watt-hours
df['power'] = df['PackCurr'] * df['PackVol']
average_power = abs(df['power'].mean())
peak_power = abs(df['power'].max())

# Calculate total watt-hours
watt_h = abs((df['PackCurr'] * df['PackVol'] * df['localtime_Diff']).sum()) / 3600
actual_ah = abs((df['PackCurr'] * df['localtime_Diff']).sum()) / 3600  # Convert seconds to hours
print("Actual Ampere-hours (Ah):------------------------------------> {:.2f}".format(actual_ah))


filtered_data_withoutDischarge = df[(df['PackCurr'] > 0) & (df['PackCurr'] < 100)]
regen_ah = abs((filtered_data_withoutDischarge['PackCurr'] * filtered_data_withoutDischarge['localtime_Diff']).sum()) / 3600  # Convert seconds to hours
average_regen_current = abs(filtered_data_withoutDischarge['PackCurr'].mean())

filtered_data_withDischarge = df[(df['PackCurr'] < 0) & (df['PackCurr'] > -150)]
average_discharge_current = abs(filtered_data_withDischarge['PackCurr'].mean())

# Calculate total distance considering power cuts
df['distance_diff'] = df['AccumulativeElapsedDist'].diff().fillna(0)
df['distance_diff'] = df['distance_diff'].apply(lambda x: x if x >= 0 else 0)   #if distance is negative then it will be considered as 0 (to handle the distance reset to 0 (which will cause negative distance difference), if negative difference, the diff will be taken as 0)
total_distance = df['distance_diff'].sum()

# Prepare analysis data for Excel output
ppt_data = {
    "Average current (A)": pack_curr_mean,
    "Average Discharge current (A)": average_discharge_current,
    "Average Regen current (A)": average_regen_current,
    "Max Discharge current (A)": abs(pack_curr_min),
    "Max Regen current (A)": pack_curr_max,
    "Average of 'ForceAtWheel' (N)": force_mean,
    "Ampere-hours Consumed (Ah)": actual_ah,
    "Regen Ampere-hours (Ah)": regen_ah,
    # "Max of 'ForceAtWheel' (N)": force_max,
    # "Min of 'ForceAtWheel' (N)": force_min,
    "Average Power (W)": average_power,
    "Peak Power (W)": peak_power,
    "Total Watt-hours (Wh)": watt_h,
    "Total Distance (km)": total_distance,
    "Efficiency (Wh/km)": watt_h / total_distance,
}

# Extract folder path and file name for output
folder_path = os.path.dirname(input_file_path)
base_filename = os.path.splitext(os.path.basename(input_file_path))[0]
excel_output_file = os.path.join(folder_path, f"analysis_{base_filename}.xlsx")

# Create a new Excel workbook and write data to it
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Analysis Results"

# Populate Excel sheet with analysis data
for i, (key, value) in enumerate(ppt_data.items(), start=1):
    ws.cell(row=i, column=1, value=key)
    ws.cell(row=i, column=2, value=value)

# Save the Excel workbook
wb.save(excel_output_file)
print("Excel output file generated at:", excel_output_file)
