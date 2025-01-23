import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import openpyxl
import os

class FileProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("File Processor")

        # File Path Input
        self.label = tk.Label(root, text="Select File (CSV or Excel):")
        self.label.pack(pady=5)

        self.path_entry = tk.Entry(root, width=50)
        self.path_entry.pack(pady=5)

        # Browse Button to select file
        self.browse_button = tk.Button(root, text="Browse", command=self.browse_file)
        self.browse_button.pack(pady=5)

        # Process Button to process the selected file
        self.process_button = tk.Button(root, text="Process File", command=self.process_file)
        self.process_button.pack(pady=10)

    def browse_file(self):
        # Allow the user to select any file
        file_path = filedialog.askopenfilename(filetypes=[("All files", "*.*"), ("CSV files", "*.csv"), ("Excel files", "*.xlsx")])
        if file_path:
            self.path_entry.delete(0, tk.END)
            self.path_entry.insert(0, file_path)

    def process_file(self):
        file_path = self.path_entry.get()

        if not file_path:
            messagebox.showerror("Error", "Please select a file to process.")
            return

        # Clean the file using the first code logic
        try:
            if file_path.endswith('.csv'):
                df = pd.read_csv(file_path)
                # print(df.head())
            elif file_path.endswith('.xlsx'):
                df = pd.read_excel(file_path)
            else:
                raise ValueError("Unsupported file format. Please select a CSV or Excel file.")

            # Remove the first 5 lines
            # df = df.iloc[4:]

            # print(df.iloc[:5, :3])
            #print header row

            print(df['DisplayName'])
            print(df['Name'])
            # Separate the initial mapping rows and the log data
            name_display_mapping = df.iloc[:153, :3]  # Adjust the row count as needed

            log_data = pd.read_csv(file_path, skiprows=154) if file_path.endswith('.csv') else pd.read_excel(file_path, skiprows=154)
            print("log_data", log_data.iloc[:1])

            

            # Create a dictionary mapping from Name to DisplayName 
            mapping_dict = pd.Series(name_display_mapping['DisplayName'].values, 
                                     index=name_display_mapping['Name']).to_dict()

            # Replace the column names in the log data using the mapping dictionary
            log_data.rename(columns=mapping_dict, inplace=True)

            # Get the directory of the input file
            input_directory = os.path.dirname(file_path)

            # Create the output file path in the same directory
            output_file_path = os.path.join(input_directory, 'modified_log_data.csv') if file_path.endswith('.csv') else os.path.join(input_directory, 'modified_log_data.xlsx')

            # Save the modified log data to the new file
            if file_path.endswith('.csv'):
                log_data.to_csv(output_file_path, index=False)
            else:
                log_data.to_excel(output_file_path, index=False)

            messagebox.showinfo("Success", f"Modified log data saved to: {output_file_path}")

            # Proceed to dyno data analysis with the cleaned file
            self.analyze_data(output_file_path)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to process the file. Error: {e}")

    def analyze_data(self, input_file_path):
        try:
            # Load the cleaned data
            if input_file_path.endswith('.csv'):
                df = pd.read_csv(input_file_path)
                # print(df.head())
            elif input_file_path.endswith('.xlsx'):
                df = pd.read_excel(input_file_path)
            else:
                raise ValueError("Unsupported file format. Please select a CSV or Excel file.")

            # Filter data based on SOC conditions
            # df = input_df[(input_df['SOC'] < 19) & (input_df['SOC'] > 2)]

            # Calculate the mean, max, and min of 'PackCurr' column
            pack_curr_mean = abs(df['PackCurr'].mean())
            pack_curr_max = df['PackCurr'].max()
            pack_curr_min = df['PackCurr'].min()

            # Calculate the mean, max, and min of 'ForceAtWheel' column
            force_mean = df['ForceAtWheel'].mean()
            force_max = df['ForceAtWheel'].max()
            force_min = df['ForceAtWheel'].min()

            # Parse the 'Date Time' column into datetime format with millisecond precision
            df['Date Time'] = pd.to_datetime(df['Date Time'], format='%d:%m:%Y %H:%M:%S:%f')

            # Calculate time difference in seconds
            df['localtime_Diff'] = df['Date Time'].diff().dt.total_seconds().fillna(0)

            # Calculate the instantaneous power and watt-hours
            df['power'] = df['PackCurr'] * df['PackVol']
            average_power = abs(df['power'].mean())
            peak_power = abs(df['power'].max())

            # Calculate total watt-hours
            watt_h = abs((df['PackCurr'] * df['PackVol'] * df['localtime_Diff']).sum()) / 3600
            actual_ah = abs((df['PackCurr'] * df['localtime_Diff']).sum()) / 3600  # Convert seconds to hours
            print("Actual Ampere-hours (Ah):------------------------------------> {:.2f}".format(actual_ah))

            filtered_data_withoutDischarge = df[(df['PackCurr'] > 0) & (df['PackCurr'] < 100)]
            regen_ah = abs((filtered_data_withoutDischarge['PackCurr'] * filtered_data_withoutDischarge['localtime_Diff']).sum()) / 3600  # Convert seconds to hours
            average_regen_current = abs(filtered_data_withoutDischarge['PackCurr'].mean())
            Regen_watt_h = abs((filtered_data_withoutDischarge['PackCurr'] * filtered_data_withoutDischarge['PackVol'] * filtered_data_withoutDischarge['localtime_Diff']).sum()) / 3600

            filtered_data_withDischarge = df[(df['PackCurr'] < 0) & (df['PackCurr'] > -150)]
            average_discharge_current = abs(filtered_data_withDischarge['PackCurr'].mean())
            Discharge_watt_h = abs((filtered_data_withDischarge['PackCurr'] * filtered_data_withDischarge['PackVol'] * filtered_data_withDischarge['localtime_Diff']).sum()) / 3600

            # Calculate total distance considering power cuts
            df['distance_diff'] = df['AccumulativeElapsedDist'].diff().fillna(0)
            df['distance_diff'] = df['distance_diff'].apply(lambda x: x if x >= 0 else 0)  # Handle negative distance
            total_distance = df['distance_diff'].sum()

            # Prepare analysis data for Excel output
            ppt_data = {
                "Average current (A)": pack_curr_mean,
                "Average Discharge current (A)": average_discharge_current,
                "Average Regen current (A)": average_regen_current,
                "Max Discharge current (A)": abs(pack_curr_min),
                "Max Regen current (A)": pack_curr_max,
                "Average of 'ForceAtWheel' (N)": force_mean,
                "Ampere-hours Consumed (Ah)": actual_ah,
                "Regen Ampere-hours (Ah)": regen_ah,
                "Discharge Energy (Wh)": Discharge_watt_h,
                "Regen Energy (Wh)": Regen_watt_h,
                "Average Power (W)": average_power,
                "Peak Power (W)": peak_power,
                "Total Watt-hours (Wh)": watt_h,
                "Total Distance (km)": total_distance,
                "Efficiency (Wh/km)": watt_h / total_distance,
                "Acceleration efficiency (Wh/km)": Discharge_watt_h/total_distance,
            }

            # Extract folder path and file name for output
            folder_path = os.path.dirname(input_file_path)
            base_filename = os.path.splitext(os.path.basename(input_file_path))[0]
            excel_output_file = os.path.join(folder_path, f"analysis_{base_filename}.xlsx")

            # Create a new Excel workbook and write data to it
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Analysis Results"

            # Populate Excel sheet with analysis data
            for i, (key, value) in enumerate(ppt_data.items(), start=1):
                ws.cell(row=i, column=1, value=key)
                ws.cell(row=i, column=2, value=value)

            # Save the Excel workbook
            wb.save(excel_output_file)
            messagebox.showinfo("Analysis Complete", f"Excel output file generated at: {excel_output_file}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to analyze data. Error: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = FileProcessorApp(root)
    root.mainloop()